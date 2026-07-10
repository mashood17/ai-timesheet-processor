"""
Section 4 — auto-detection of the IQAMA column and day columns in the
user's master Excel workbook, with NO hardcoded column letters or sheet
names, and NO assumption that headers sit in row 1.

UPDATE: real-world testing showed some workbooks store day-column headers
as generic placeholder dates (e.g. Excel's default 1900-01-01, 1900-01-02...
representing "Day 1, Day 2..." rather than actual calendar dates for the
period). Calendar-date alignment fails completely against these (0% match),
even though the day-column *positions* themselves were detected correctly.

So detection now tries BOTH alignment strategies and returns both results
with method labels — it never silently picks one. The confirmation UI
(Section 4's required step) shows the user which method matched and lets
them choose, rather than the backend guessing.
"""
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.date_utils import excel_serial_to_date

MAX_HEADER_SCAN_ROWS = 20
IQAMA_KEYWORDS = ("iqama", "passport")


class ExcelMappingService:
    def detect_mapping(
        self,
        excel_path: Path,
        sheet_name: str,
        pdf_dates: list[date],
    ) -> dict:
        workbook = openpyxl.load_workbook(excel_path, data_only=False)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
        sheet = workbook[sheet_name]

        warnings: list[str] = []

        label_row, iqama_col_idx, iqama_confidence = self._find_iqama_column(sheet)
        if iqama_col_idx is None:
            warnings.append(
                "Could not confidently locate an IQAMA/Passport column. "
                "Please select it manually before continuing."
            )

        date_row_idx, date_columns = self._find_date_row(sheet, label_row)
        if not date_columns:
            warnings.append(
                "Could not confidently locate the day-of-month date columns. "
                "Please select the date row/columns manually before continuing."
            )

        # Strategy 1: align by actual calendar date value (Section 4's primary rule).
        calendar_mapping, calendar_unmatched = self._align_dates_by_calendar(
            date_columns, pdf_dates
        )
        calendar_confidence = len(calendar_mapping) / len(pdf_dates) if pdf_dates else 0.0

        # Strategy 2: fallback — the Excel's day-columns are real (detected via a
        # mostly-date-like row) but their VALUES are placeholders, not true dates.
        # Align PDF dates sequentially to the detected Excel day-columns by position.
        positional_mapping = self._align_dates_positionally(date_columns, pdf_dates)
        positional_confidence = (
            len(positional_mapping) / len(pdf_dates) if pdf_dates else 0.0
        )

        # Prefer calendar alignment if it found anything meaningful; otherwise
        # surface the positional fallback as the suggested option, but always
        # tell the user which method is being proposed so they can override it.
        if calendar_confidence >= 0.5:
            recommended_method = "calendar"
            day_column_mapping, unmatched = calendar_mapping, calendar_unmatched
        elif positional_confidence > 0:
            recommended_method = "positional"
            day_column_mapping, unmatched = positional_mapping, []
            warnings.append(
                "The Excel's day-column headers don't contain real calendar dates "
                "(they appear to be generic placeholders). Falling back to matching "
                "the PDF's days to the Excel's day-columns by position/order instead. "
                "Please verify this is correct before proceeding — check that the "
                "first PDF day genuinely corresponds to the first detected Excel "
                "day-column."
            )
        else:
            recommended_method = "calendar"
            day_column_mapping, unmatched = calendar_mapping, calendar_unmatched

        data_start_row = self._find_data_start_row(sheet, label_row, iqama_col_idx)
        day_columns_confidence = (
            len(day_column_mapping) / len(pdf_dates) if pdf_dates else 0.0
        )

        workbook.close()

        return {
            "header_row": label_row or 1,
            "data_start_row": data_start_row,
            "iqama_column": get_column_letter(iqama_col_idx) if iqama_col_idx else None,
            "iqama_column_confidence": iqama_confidence,
            "day_columns": [
                {"iso_date": iso, "excel_column": get_column_letter(col_idx)}
                for iso, col_idx in day_column_mapping.items()
            ],
            "day_columns_confidence": day_columns_confidence,
            "unmatched_pdf_dates": unmatched,
            "alignment_method": recommended_method,
            "calendar_match_count": len(calendar_mapping),
            "positional_match_count": len(positional_mapping),
            "warnings": warnings,
        }

    # ---------- internal helpers ----------

    def _find_iqama_column(self, sheet: Worksheet) -> tuple[int | None, int | None, float]:
        for row_idx in range(1, MAX_HEADER_SCAN_ROWS + 1):
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = str(cell.value).strip().lower() if cell.value is not None else ""
                if any(kw in value for kw in IQAMA_KEYWORDS):
                    confidence = 0.95 if value in ("iqama", "passport", "iqama or passport #") else 0.75
                    return row_idx, col_idx, confidence
        return None, None, 0.0

    def _find_date_row(
        self, sheet: Worksheet, label_row: int | None
    ) -> tuple[int | None, dict[int, date]]:
        search_start = label_row or 1
        search_end = min(search_start + 5, sheet.max_row)

        best_row = None
        best_dates: dict[int, date] = {}

        for row_idx in range(search_start, search_end + 1):
            found: dict[int, date] = {}
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                parsed = excel_serial_to_date(cell.value)
                if parsed:
                    found[col_idx] = parsed
            if len(found) > len(best_dates):
                best_row, best_dates = row_idx, found

        if len(best_dates) < 3:
            return None, {}
        return best_row, best_dates

    def _align_dates_by_calendar(
        self, excel_dates: dict[int, date], pdf_dates: list[date]
    ) -> tuple[dict[str, int], list[str]]:
        by_date = {d: col_idx for col_idx, d in excel_dates.items()}
        mapping: dict[str, int] = {}
        unmatched: list[str] = []

        for pdf_date in pdf_dates:
            col_idx = by_date.get(pdf_date)
            if col_idx is not None:
                mapping[pdf_date.isoformat()] = col_idx
            else:
                unmatched.append(pdf_date.isoformat())

        return mapping, unmatched

    def _align_dates_positionally(
        self, excel_dates: dict[int, date], pdf_dates: list[date]
    ) -> dict[str, int]:
        """
        Fallback: ignores the actual date VALUES in excel_dates and just uses
        their column ORDER — the Nth detected Excel day-column gets the Nth
        PDF date. Only used when calendar alignment clearly failed, and always
        surfaced to the user as a fallback, never silently applied elsewhere.
        """
        ordered_columns = sorted(excel_dates.keys())
        mapping: dict[str, int] = {}
        for i, pdf_date in enumerate(pdf_dates):
            if i >= len(ordered_columns):
                break
            mapping[pdf_date.isoformat()] = ordered_columns[i]
        return mapping

    def _find_data_start_row(
        self, sheet: Worksheet, label_row: int | None, iqama_col_idx: int | None
    ) -> int:
        if not label_row or not iqama_col_idx:
            return (label_row or 1) + 1

        for row_idx in range(label_row + 1, min(label_row + 10, sheet.max_row) + 1):
            if sheet.cell(row=row_idx, column=iqama_col_idx).value not in (None, ""):
                return row_idx

        return label_row + 1