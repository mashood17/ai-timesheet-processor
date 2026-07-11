"""
Section 5, steps 11-12 + Section 4: matches PDF employee rows against the
master Excel's IQAMA column, and separately flags duplicate IQAMAs found
within the PDF itself (a data-quality problem in the source document,
independent of matching).
"""
import re
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string


def normalize_id(value) -> str:
    """
    Normalizes an IQAMA/passport value for comparison. Strips whitespace
    (including invisible/non-breaking unicode whitespace), keeps only
    alphanumerics, uppercases for case-insensitive comparison.

    Deliberately does NOT attempt fuzzy correction of common OCR confusions
    — per explicit requirement, this must never create a false match. See
    MatchingService.find_possible_match for the safe, human-confirmed
    alternative.
    """
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", "", text)
    return re.sub(r"[^0-9A-Za-z]", "", text).upper()


class MatchingService:
    def build_iqama_index(
        self,
        excel_path: Path,
        sheet_name: str,
        iqama_column: str,
        data_start_row: int,
    ) -> dict[str, int]:
        workbook = openpyxl.load_workbook(excel_path, data_only=False, read_only=True)
        sheet = workbook[sheet_name]
        col_idx = column_index_from_string(iqama_column)

        index: dict[str, int] = {}
        for row_idx in range(data_start_row, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            normalized = normalize_id(cell_value)
            if normalized and normalized not in index:
                index[normalized] = row_idx

        workbook.close()
        return index

    def find_duplicate_pdf_iqamas(self, pdf_iqamas: list[str]) -> dict[str, int]:
        normalized = [normalize_id(v) for v in pdf_iqamas if v]
        counts = Counter(normalized)
        return {iqama: count for iqama, count in counts.items() if count > 1}

    def find_possible_match(
        self, iqama: str, index: dict[str, int], max_distance: int = 2, min_length: int = 6
    ) -> str | None:
        """
        BUGFIX: this was previously a stray module-level function, called
        incorrectly as an instance method everywhere it was used — an
        AttributeError that crashed both /api/process and /api/process/
        accept-match with a 500. Now a proper method on this class.

        Suggests a close-but-not-exact IQAMA match for human review —
        NEVER auto-applied, NEVER used for matching itself. Only returns a
        suggestion when there is exactly ONE unambiguous close candidate;
        if two candidates are equally close, returns None rather than
        guessing between them. Pure generic edit-distance comparison — no
        assumptions about which specific digits get confused.
        """
        if not iqama or len(iqama) < min_length:
            return None

        best_candidate: str | None = None
        best_distance = max_distance + 1
        second_best_distance = max_distance + 1

        for candidate in index:
            if abs(len(candidate) - len(iqama)) > 1:
                continue
            distance = self._levenshtein(iqama, candidate)
            if distance < best_distance:
                second_best_distance = best_distance
                best_distance = distance
                best_candidate = candidate
            elif distance < second_best_distance:
                second_best_distance = distance

        if best_candidate and best_distance <= max_distance and best_distance < second_best_distance:
            return best_candidate
        return None

    @staticmethod
    def _levenshtein(a: str, b: str) -> int:
        if len(a) < len(b):
            a, b = b, a
        previous_row = list(range(len(b) + 1))
        for i, char_a in enumerate(a, start=1):
            current_row = [i]
            for j, char_b in enumerate(b, start=1):
                cost = 0 if char_a == char_b else 1
                current_row.append(min(
                    previous_row[j] + 1,
                    current_row[j - 1] + 1,
                    previous_row[j - 1] + cost,
                ))
            previous_row = current_row
        return previous_row[-1]