"""
Section 5 step 16 — generates the validation/error report as a readable
.xlsx (not a raw text dump, per Section 11), covering:
  - unmatched IQAMAs
  - duplicate IQAMAs found in the PDF
  - manually-corrected cells (audit trail: who/what was overridden and why)
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from app.models.process_models import (
    CorrectedCell,
    DuplicateEntry,
    EmployeeProcessResult,
    UnmatchedEntry,
)

HEADER_FILL = PatternFill(start_color="2D3142", end_color="2D3142", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


class ReportService:
    def generate_report(
        self,
        output_path: Path,
        unmatched: list[UnmatchedEntry],
        duplicates: list[DuplicateEntry],
        corrections: list[CorrectedCell],
        results: list[EmployeeProcessResult],
    ) -> None:
        workbook = openpyxl.Workbook()

        self._write_unmatched_sheet(workbook.active, unmatched)
        self._write_duplicates_sheet(workbook.create_sheet("Duplicate IQAMAs"), duplicates)
        self._write_corrections_sheet(
            workbook.create_sheet("Manual Corrections"), corrections, results
        )

        workbook.save(output_path)

    def _write_unmatched_sheet(self, sheet, unmatched: list[UnmatchedEntry]) -> None:
        sheet.title = "Unmatched IQAMAs"
        headers = ["IQAMA / Passport", "Employee Name", "Reason"]
        self._write_header(sheet, headers)
        for row_idx, entry in enumerate(unmatched, start=2):
            sheet.cell(row=row_idx, column=1, value=entry.iqama_or_passport)
            sheet.cell(row=row_idx, column=2, value=entry.employee_name)
            sheet.cell(row=row_idx, column=3, value=entry.reason)
        self._autosize(sheet, len(headers))

    def _write_duplicates_sheet(self, sheet, duplicates: list[DuplicateEntry]) -> None:
        headers = ["IQAMA / Passport", "Occurrences", "Employee Name(s) Found"]
        self._write_header(sheet, headers)
        for row_idx, entry in enumerate(duplicates, start=2):
            sheet.cell(row=row_idx, column=1, value=entry.iqama_or_passport)
            sheet.cell(row=row_idx, column=2, value=entry.occurrences)
            sheet.cell(row=row_idx, column=3, value=", ".join(entry.employee_names))
        self._autosize(sheet, len(headers))

    def _write_corrections_sheet(
        self,
        sheet,
        corrections: list[CorrectedCell],
        results: list[EmployeeProcessResult],
    ) -> None:
        """
        Audit trail: shows what the OCR originally read vs. what the user
        confirmed, so a payroll auditor can see every manual override.
        """
        headers = ["IQAMA / Passport", "Employee Name", "Date", "OCR Original Value", "Corrected Value"]
        self._write_header(sheet, headers)

        # Build a lookup of original OCR values for the audit trail.
        original_lookup: dict[tuple[str, str], str] = {}
        name_lookup: dict[str, str | None] = {}
        for employee in results:
            key_iqama = (employee.iqama_or_passport or "").strip().upper()
            name_lookup[key_iqama] = employee.employee_name
            for cell in employee.day_cells:
                original_lookup[(key_iqama, cell.iso_date)] = cell.value

        for row_idx, correction in enumerate(corrections, start=2):
            key_iqama = (correction.iqama_or_passport or "").strip().upper()
            sheet.cell(row=row_idx, column=1, value=correction.iqama_or_passport)
            sheet.cell(row=row_idx, column=2, value=name_lookup.get(key_iqama))
            sheet.cell(row=row_idx, column=3, value=correction.iso_date)
            sheet.cell(row=row_idx, column=4, value=original_lookup.get((key_iqama, correction.iso_date)))
            sheet.cell(row=row_idx, column=5, value=correction.corrected_value)

        self._autosize(sheet, len(headers))

    def _write_header(self, sheet, headers: list[str]) -> None:
        for col_idx, text in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=text)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

    def _autosize(self, sheet, num_columns: int) -> None:
        for col_idx in range(1, num_columns + 1):
            letter = sheet.cell(row=1, column=col_idx).column_letter
            max_len = max(
                (len(str(sheet.cell(row=r, column=col_idx).value or "")) for r in range(1, sheet.max_row + 1)),
                default=10,
            )
            sheet.column_dimensions[letter].width = min(max(max_len + 2, 12), 50)